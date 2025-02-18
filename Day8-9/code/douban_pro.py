import requests
import os
from lxml import etree
import time
import pandas as pd
from openpyxl.drawing.image import Image
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

current_dir = os.path.dirname(os.path.abspath(__file__))
save_path = os.path.join(current_dir, 'douban_250.xlsx')

def fetch_250():
    headers = {
        "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/133.0.0.0 Safari/537.36',
        "Cookie": 'bid=JgE1tB35B_Q; ap_v=0,6.0; __utma=30149280.1943493433.1739439174.1739439174.1739439174.1; __utmc=30149280; __utmz=30149280.1739439174.1.1.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided); push_noty_num=0; push_doumail_num=0; __utmt=1; __utmv=30149280.28696; dbcl2="286964664:xJc/sCfBNYk"; ck=bhE4; frodotk_db="843a16c65e232c8a8cae44531a69859e"; ll="108288"; __utmb=30149280.19.6.1739440854876; _pk_ref.100001.4cf6=%5B%22%22%2C%22%22%2C1739440860%2C%22https%3A%2F%2Fwww.douban.com%2Flink2%2F%3Furl%3Dhttp%3A%2F%2Fwww.douban.com%2Fmovie%2Ftop250%22%5D; _pk_id.100001.4cf6=c3a5b68024ade6c4.1739440860.; _pk_ses.100001.4cf6=1; __utma=223695111.814996443.1739440860.1739440860.1739440860.1; __utmb=223695111.0.10.1739440860; __utmc=223695111; __utmz=223695111.1739440860.1.1.utmcsr=douban.com|utmccn=(referral)|utmcmd=referral|utmcct=/link2/'
    }
    movies = []
    for i in range(0, 250, 25):
        url = f'https://movie.douban.com/top250?start={i}'
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            response.encoding = 'utf-8'
            html = response.text

            tree = etree.HTML(html)

            items = tree.xpath('//div[@class = "item"]')
            for item in items:
                name = item.xpath('.//span[@class = "title"][1]/text()')[0]
                director = item.xpath('.//p[1]/text()')[0].strip().split('导演: ')[1].split(' ')[0]
                pub_time = item.xpath('.//p[1]/text()')[1].strip().split('/')[0].strip()
                score = item.xpath('.//span[@class="rating_num"]/text()')[0]
                quote = item.xpath('.//span[@class="inq"]/text()')
                quote = quote[0] if quote else ''
                img_link = item.xpath('.//img/@src')[0]
                movies.append({
                    "标题": name,
                    "导演": director,
                    "评分": score,
                    "上映时间": pub_time,
                    "经典台词": quote,
                    "封面链接": img_link
                })
        except requests.RequestException as e:
            print(f'请求出错: {e}')
        except IndexError as e:
            print(f'解析出错: {e}')
        time.sleep(1)
    return movies

def clean_data(movies):
    df = pd.DataFrame(movies)
    df["评分"] = df["评分"].astype(float)
    df['经典台词'] = df['经典台词'].fillna('暂无台词')
    return df

def save_to_excel(df):
    with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="电影数据")

        worksheet = writer.sheets["电影数据"]
        worksheet.column_dimensions["A"].width = 30
        worksheet.column_dimensions["D"].width = 10
        worksheet.column_dimensions["B"].width = 35
        worksheet.column_dimensions["C"].width = 10
        worksheet.column_dimensions["E"].width = 90
        worksheet.column_dimensions["F"].width = 90

    print("Excel保存成功")

def insert_image(df, filename):
    wb = load_workbook(filename)
    ws = wb.active

    for idx, row in enumerate(df.itertuples(), start=2):
        img_url = row.封面链接
        try:
            img_data = requests.get(img_url).content
            img = Image(io.BytesIO(img_data))
            img.width, img.height = 100, 150

            col_letter = get_column_letter(7)  # G
            cell = f'{col_letter}{idx}'
            ws.column_dimensions[col_letter].width = 20  
            ws.row_dimensions[idx].height = 120  

            ws.add_image(img, cell)
            print(f"第{idx}张图片保存成功")

        except requests.RequestException as e:
            print(f"下载图片出错: {e}")

    wb.save(filename)

movies = fetch_250()
df = clean_data(movies)
save_to_excel(df)
insert_image(df, save_path)